#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Part 1：计算当前使用的 8 个强相关 CIF 结构描述符。

在项目根目录运行：

    python3 因果推断/part1.py

或进入脚本所在目录后运行：

    python3 part1.py

默认计算并输出全部 8 个强相关描述符；也可以由使用者指定只输出其中一部分：

    python3 part1.py --descriptors A2 A4 A9
    python3 part1.py --descriptors 1 2 3
    python3 part1.py --descriptors 局域宽松因子_A2,Na多面体畸变均值
    python3 part1.py --list-descriptors

如果需要把“所选描述符全部达到快导体筛选阈值”的 CIF 移动到新文件夹，
需要手动启用：

    python3 part1.py --descriptors 1 2 3 --move-fast-cifs

脚本默认读取 `data/快慢离子导体数据集_107.xlsx`，用于保留材料顺序，
并按 `合并编号` 前缀匹配 CIF 文件。也可以在不读取 Excel 的情况下，
直接处理一个 CIF 文件夹。

以下经验阈值采用的快/慢导体标签：
    快导体：离子电导率 >= 0.1 mS/cm
    慢导体：离子电导率 < 0.1 mS/cm

本脚本使用的描述符筛选阈值：
    1. 局域宽松因子_A2 [即 局域宽松因子]
       - 慢导体富集区：A2 < 1.09
       - 快导体富集区：A2 >= 1.19
       - 当前 103 样本平衡准确率经验筛选阈值：A2 >= 1.118107
    2. Na多面体畸变均值
       - 经验筛选阈值：>= 0.043210
    3. Na位点连通因子_NaNa综合
       - 经验筛选阈值：>= 0.482324
    4. 局域_连通协同因子_A2xNaNa
       - 经验筛选阈值：>= 0.577777
       - 机制导向快导体富集规则：A2 高且 NaNa 连通强
5. 畸变乘键长比_A5 [即 畸变加权宽松比]
        - 经验筛选阈值：>= 0.040056
     6. 畸变乘最长键长_A4 [即 畸变加权最长键]
        - 经验筛选阈值：>= 0.126580
     7. 畸变除Na浓度_A9 [即 畸变密度]
       - 经验筛选阈值：>= 2.314122
    8. Na-S_avg_bond_norm [即 硫化物键长归一化]
       - 仅适用于硫化物族内的经验范围：>= 1.039905，约 1.04-1.05

重要说明：这些阈值是从当前 103 个 Na-only 样本中得到的筛选阈值。
它们不是普适物理常数。用于独立高通量数据集前，应重新校准。

输出文件：
    part1_descriptors.csv
    part1_descriptors.json
    part1_issues.csv

8 个输出描述符：
    - 局域宽松因子_A2 = Na-X最长键长_A / Na-X目标键长中心_A
    - Na多面体畸变均值
    - Na位点连通因子_NaNa综合  # 直观名: Na位点连通因子
    - 局域_连通协同因子_A2xNaNa  # 直观名: 局域-连通协同因子
    - 畸变乘键长比_A5 = Na多面体畸变均值 * 局域宽松因子_A2
    - 畸变乘最长键长_A4 = Na多面体畸变均值 * Na-X最长键长_A
    - 畸变除Na浓度_A9 = Na多面体畸变均值 / Na浓度_A-3
    - Na-S_avg_bond_norm（硫化物族内辅助描述符）

有意排除：
    - Zeo++ 描述符
    - SoftBV / BVSE 描述符
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, pstdev
from typing import Any

import numpy as np
from openpyxl import load_workbook
from pymatgen.io.cif import CifParser
from pymatgen.symmetry.analyzer import SpacegroupAnalyzer

try:
    from scipy.spatial import ConvexHull, QhullError
except Exception:  # 测试覆盖不到的可选依赖兜底
    ConvexHull = None
    QhullError = Exception


MOBILE = "Na"
DEFAULT_WORKBOOK = "data/快慢离子导体数据集_107.xlsx"
DEFAULT_CIF_DIR = "cif"
DEFAULT_SHEET = "汇报主表"
FAST_CONDUCTIVITY_THRESHOLD_MS_CM = 0.1
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

# 默认输出哪些强相关描述符。
# 保留 ["all"] 表示全部 8 个；也可以改成 ["A2", "A4", "A9"] 这样的短名列表。
# 运行命令中传入 --descriptors 时，会覆盖这里的默认设置。
DEFAULT_SELECTED_DESCRIPTORS = ["all"]

# 手动启用 --move-fast-cifs 时，满足阈值的 CIF 会被移动到这个文件夹。
DEFAULT_FAST_CIF_OUTPUT_DIR = "阈值筛选快离子导体CIF"

# 阈值状态说明：
#   阶段报告支持：已有阶段报告明确支持的阈值。
#   当前数据经验阈值：在当前 103 样本上按平衡准确率得到的经验阈值。
#   族内专用：只适用于指定化学体系或材料族内。
DESCRIPTOR_THRESHOLDS: dict[str, dict[str, Any]] = {
    "局域宽松因子_A2": {
        "slow_enriched_lt": 1.09,
        "fast_enriched_ge": 1.19,
        "screening_ge": 1.118107,
        "status": "阶段报告支持 + 当前数据经验阈值",
        "note": "局域宽松因子: A2 < 1.09 为慢导体富集区；A2 >= 1.19 为快导体富集区；1.118107 是当前标签下平衡准确率最优阈值。",
    },
    "Na多面体畸变均值": {
        "screening_ge": 0.043210,
        "status": "当前数据经验阈值",
        "note": "跨体系相对稳定的辅助描述符；该阈值仅用于当前数据筛选。",
    },
    "Na位点连通因子_NaNa综合": {
        "screening_ge": 0.482324,
        "status": "当前数据经验阈值",
        "note": "Na位点连通因子: 由 Na 平均邻居数和最大 Na 连通分量占比的百分位秩综合得到。",
    },
    "局域_连通协同因子_A2xNaNa": {
        "screening_ge": 0.577777,
        "status": "当前数据经验阈值",
        "note": "局域-连通协同因子: 机制假说描述符，不是已证明的因果规则。",
    },
    "畸变乘键长比_A5": {
        "screening_ge": 0.040056,
        "status": "当前数据经验阈值",
        "note": "畸变加权宽松比: Na 多面体畸变均值与 A2 的乘积。",
    },
    "畸变乘最长键长_A4": {
        "screening_ge": 0.126580,
        "status": "当前数据经验阈值",
        "note": "畸变加权最长键: Na 多面体畸变均值与 Na-X 最长键长的乘积。",
    },
    "畸变除Na浓度_A9": {
        "screening_ge": 2.314122,
        "status": "当前数据经验阈值",
        "note": "畸变密度: Na 多面体畸变均值除以 Na 浓度；该描述符可能受尺度影响。",
    },
    "Na-S_avg_bond_norm": {
        "screening_ge": 1.039905,
        "status": "族内专用",
        "note": "硫化物键长归一化(Na-S): 仅用于硫化物；近似实用范围为 1.04-1.05。",
    },
}

DESCRIPTOR_COLUMNS = list(DESCRIPTOR_THRESHOLDS.keys())

# 允许用短名选择描述符，避免每次输入完整中文列名。
DESCRIPTOR_ALIASES: dict[str, str] = {
    "a2": "局域宽松因子_A2",  # 直观名: 局域宽松因子
    "局域宽松因子": "局域宽松因子_A2",  # 直观名: 局域宽松因子
    "键长比": "局域宽松因子_A2",  # 直观名: 局域宽松因子
    "distortion": "Na多面体畸变均值",
    "na_poly_distortion": "Na多面体畸变均值",
    "畸变": "Na多面体畸变均值",
    "畸变均值": "Na多面体畸变均值",
    "nana": "Na位点连通因子_NaNa综合",  # 直观名: Na位点连通因子
    "na-na": "Na位点连通因子_NaNa综合",  # 直观名: Na位点连通因子
    "连通因子": "Na位点连通因子_NaNa综合",  # 直观名: Na位点连通因子
    "a2xnana": "局域_连通协同因子_A2xNaNa",  # 直观名: 局域-连通协同因子
    "a2*nana": "局域_连通协同因子_A2xNaNa",  # 直观名: 局域-连通协同因子
    "协同因子": "局域_连通协同因子_A2xNaNa",  # 直观名: 局域-连通协同因子
    "a5": "畸变乘键长比_A5",  # 直观名: 畸变加权宽松比
    "畸变乘键长比": "畸变乘键长比_A5",  # 直观名: 畸变加权宽松比
    "a4": "畸变乘最长键长_A4",  # 直观名: 畸变加权最长键
    "畸变乘最长键长": "畸变乘最长键长_A4",  # 直观名: 畸变加权最长键
    "a9": "畸变除Na浓度_A9",  # 直观名: 畸变密度
    "畸变除na浓度": "畸变除Na浓度_A9",  # 直观名: 畸变密度
    "畸变除钠浓度": "畸变除Na浓度_A9",  # 直观名: 畸变密度
    "na-s": "Na-S_avg_bond_norm",  # 直观名: 硫化物键长归一化
    "nas": "Na-S_avg_bond_norm",  # 直观名: 硫化物键长归一化
    "na_s": "Na-S_avg_bond_norm",  # 直观名: 硫化物键长归一化
    "硫化物键长归一化": "Na-S_avg_bond_norm",  # 直观名: 硫化物键长归一化
}

ANION_ELEMENTS = {"O", "S", "Se", "F", "Cl", "Br", "I", "N", "H"}

# 沿用前期锁定流程中的经典有效离子半径表。
# 输出列名使用“有效离子半径”，不使用英文人名命名。
NA_EFFECTIVE_RADII_A: dict[int, float] = {
    4: 0.99,
    5: 1.00,
    6: 1.02,
    7: 1.12,
    8: 1.18,
    9: 1.24,
    12: 1.39,
}
NA_FALLBACK_CN = 6

ANION_EFFECTIVE_RADII_A: dict[str, float | None] = {
    "O": 1.40,
    "S": 1.84,
    "Se": 1.98,
    "F": 1.33,
    "Cl": 1.81,
    "Br": 1.96,
    "I": 2.20,
    "H": 1.40,
    "N": None,
}


BASE_OUTPUT_COLUMNS = [
    "合并编号",
    "CIF文件",
    "体系分类",
    "材料/结构名",
    "电导率_mS_cm-1",
    "log10电导率",
    "Na邻近阴离子类型",
]

STATUS_OUTPUT_COLUMNS = [
    "解析状态",
    "问题数量",
    "问题",
]

OUTPUT_COLUMNS = BASE_OUTPUT_COLUMNS + DESCRIPTOR_COLUMNS + STATUS_OUTPUT_COLUMNS


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def round_or_none(value: float | None, digits: int = 4) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return round(float(value), digits)


def safe_float(value: Any) -> float | None:
    value = clean(value)
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return f


def log10_or_none(value: Any) -> float | None:
    f = safe_float(value)
    if f is None or f <= 0:
        return None
    return round(math.log10(f), 6)


def fmt_range(values: list[float], digits: int = 3) -> str | None:
    finite = [float(v) for v in values if v is not None and not math.isnan(float(v))]
    if not finite:
        return None
    return f"{min(finite):.{digits}f}-{max(finite):.{digits}f}"


def site_species_dict(site) -> dict[str, float]:
    out: dict[str, float] = defaultdict(float)
    for el, occ in site.species.items():
        out[str(getattr(el, "symbol", str(el)))] += float(occ)
    return dict(out)


def major_symbol(site) -> str:
    species = site_species_dict(site)
    return max(species.items(), key=lambda kv: kv[1])[0] if species else ""


def site_occ(site, symbol: str) -> float:
    return site_species_dict(site).get(symbol, 0.0)


def is_disordered_or_partial(site) -> bool:
    species = site_species_dict(site)
    total = sum(species.values())
    if len(species) != 1:
        return True
    if abs(total - 1.0) > 1e-3:
        return True
    return any(abs(v - 1.0) > 1e-3 for v in species.values())


def anion_cutoff(anion_symbols: set[str]) -> float:
    cutoffs = {
        "O": 3.20,
        "F": 3.20,
        "N": 3.35,
        "S": 3.85,
        "Cl": 3.85,
        "H": 3.20,
        "Se": 4.05,
        "Br": 4.05,
        "I": 4.35,
    }
    return max((cutoffs.get(sym, 4.0) for sym in anion_symbols), default=4.0)


def unpack_neighbor(item: Any, center_coords: np.ndarray) -> tuple[Any, float, int | None, np.ndarray]:
    if isinstance(item, tuple):
        site = item[0]
        dist = float(item[1])
        idx = int(item[2]) if len(item) >= 3 and item[2] is not None else None
        return site, dist, idx, np.array(site.coords, dtype=float)
    site = item
    dist = getattr(item, "nn_distance", None)
    if dist is None:
        dist = float(np.linalg.norm(np.array(site.coords, dtype=float) - center_coords))
    idx = getattr(item, "index", None)
    return site, float(dist), idx, np.array(site.coords, dtype=float)


def shell_neighbors(structure, center_index: int, anion_symbols: set[str]) -> list[dict[str, Any]]:
    """按参考脚本的简化规则提取第一配位壳层 Na-X 近邻。"""
    center = structure[center_index]
    raw = structure.get_sites_in_sphere(
        center.coords,
        anion_cutoff(anion_symbols),
        include_index=True,
        include_image=True,
    )
    center_coords = np.array(center.coords, dtype=float)
    neighbors: list[dict[str, Any]] = []
    for item in raw:
        site, dist, idx, coords = unpack_neighbor(item, center_coords)
        if idx == center_index and dist < 1e-6:
            continue
        sym = major_symbol(site)
        if sym in anion_symbols:
            neighbors.append({"symbol": sym, "distance": dist, "coords": coords, "index": idx})
    neighbors.sort(key=lambda x: x["distance"])
    if not neighbors:
        return []
    first = neighbors[0]["distance"]
    kept = [n for n in neighbors if n["distance"] <= first + 0.70]
    if len(kept) <= 3 and len(neighbors) > len(kept):
        kept = neighbors[: min(4, len(neighbors))]
    return kept


def mode_int(values: list[int]) -> int | None:
    if not values:
        return None
    return sorted(Counter(values).items(), key=lambda kv: (-kv[1], kv[0]))[0][0]


def counter_text(counter: Counter) -> str | None:
    if not counter:
        return None
    return "; ".join(f"{k}:{v}" for k, v in sorted(counter.items(), key=lambda kv: str(kv[0])))


def convex_volume(coords: list[np.ndarray]) -> float | None:
    if len(coords) < 4 or ConvexHull is None:
        return None
    try:
        return float(ConvexHull(np.array(coords, dtype=float)).volume)
    except (QhullError, ValueError):
        return None


def cn_to_polyhedron_name(cn: int) -> str:
    return {
        2: "linear",
        3: "trigonal-planar",
        4: "tetrahedron",
        5: "trig-bipyramid/sq-pyramid",
        6: "octahedron",
        7: "capped-octahedron",
        8: "cube/sq-antiprism",
        9: "tricapped-trig-prism",
        10: "bicapped-sq-antiprism",
        12: "cuboctahedron",
    }.get(cn, f"CN{cn}")


def effective_na_radius(cn: int | None) -> tuple[float | None, str | None]:
    if cn in NA_EFFECTIVE_RADII_A:
        return NA_EFFECTIVE_RADII_A[int(cn)], f"Na+(CN={int(cn)}, 经典有效离子半径表)"
    fallback = NA_EFFECTIVE_RADII_A[NA_FALLBACK_CN]
    if cn is None:
        return fallback, f"Na+(CN={NA_FALLBACK_CN}, 经典有效离子半径表默认值; CN缺失)"
    return fallback, f"Na+(CN={NA_FALLBACK_CN}, 经典有效离子半径表默认值; 观测CN={cn}未列入)"


def effective_anion_radius(anion_symbols: set[str]) -> tuple[float | None, str | None]:
    if not anion_symbols:
        return None, None
    values: list[tuple[str, float]] = []
    missing: list[str] = []
    for sym in sorted(anion_symbols):
        radius = ANION_EFFECTIVE_RADII_A.get(sym)
        if radius is None:
            missing.append(sym)
        else:
            values.append((sym, radius))
    if missing:
        return None, f"{'/'.join(missing)} 未锁定有效离子半径"
    if not values:
        return None, None
    if len(values) == 1:
        sym, radius = values[0]
        return radius, f"{sym}({radius:.2f} A, 经典有效离子半径表)"
    avg = sum(radius for _sym, radius in values) / len(values)
    source = "多阴离子简单平均: " + ", ".join(f"{sym}({radius:.2f} A)" for sym, radius in values)
    return avg, source


def nearest_periodic_distances(structure, indices_a: list[int], indices_b: list[int], exclude_same: bool) -> list[float]:
    distances = []
    for ia in indices_a:
        best = None
        for ib in indices_b:
            if exclude_same and ia == ib:
                continue
            d = float(structure.get_distance(ia, ib))
            if best is None or d < best:
                best = d
        if best is not None:
            distances.append(best)
    return distances


def build_na_network(structure, na_indices: list[int], cutoff: float) -> dict[str, Any]:
    n = len(na_indices)
    if n < 2:
        return {
            "Na-Na最近距离_A": None,
            "Na-Na平均最近邻距离_A": None,
            "Na-Na距离标准差_A": None,
            "每个Na平均Na邻居数": None,
            "Na邻居数最大值": None,
            "Na网络连通分量数": None,
            "最大Na连通分量占比": None,
            "Na网络维度估计": None,
        }

    all_distances: list[float] = []
    nearest: list[float] = []
    neighbors: list[set[int]] = [set() for _ in range(n)]
    for i in range(n):
        local: list[tuple[int, float]] = []
        for j in range(n):
            if i == j:
                continue
            d = float(structure.get_distance(na_indices[i], na_indices[j]))
            local.append((j, d))
            all_distances.append(d)
            if d <= cutoff:
                neighbors[i].add(j)
        if local:
            nearest.append(min(d for _j, d in local))

    seen: set[int] = set()
    components: list[set[int]] = []
    for start in range(n):
        if start in seen:
            continue
        stack = [start]
        comp: set[int] = set()
        seen.add(start)
        while stack:
            cur = stack.pop()
            comp.add(cur)
            for nxt in neighbors[cur]:
                if nxt not in seen:
                    seen.add(nxt)
                    stack.append(nxt)
        components.append(comp)

    largest = max((len(c) for c in components), default=0)
    largest_frac = largest / n if n else None
    counts = [len(x) for x in neighbors]

    coords = np.array([structure[i].frac_coords for i in na_indices], dtype=float)
    spans = []
    for ax in range(3):
        cs = np.sort(coords[:, ax])
        gaps = np.diff(cs)
        wrap_gap = (cs[0] + 1.0) - cs[-1]
        max_gap = max(gaps.max() if len(gaps) else 0.0, wrap_gap)
        spans.append(1.0 - max_gap)
    s_sorted = sorted(spans, reverse=True)
    if largest_frac is not None and largest_frac >= 0.8 and s_sorted[2] > 0.55:
        dim = "3D倾向"
    elif largest_frac is not None and largest_frac >= 0.5 and s_sorted[1] > 0.55:
        dim = "2D倾向"
    elif largest_frac is not None and largest_frac >= 0.3 and s_sorted[0] > 0.55:
        dim = "1D倾向"
    else:
        dim = "低连通/局域团簇"

    return {
        "Na-Na最近距离_A": round_or_none(min(all_distances) if all_distances else None, 4),
        "Na-Na平均最近邻距离_A": round_or_none(mean(nearest) if nearest else None, 4),
        "Na-Na距离标准差_A": round_or_none(pstdev(all_distances) if len(all_distances) > 1 else None, 4),
        "每个Na平均Na邻居数": round_or_none(mean(counts) if counts else None, 3),
        "Na邻居数最大值": max(counts) if counts else None,
        "Na网络连通分量数": len(components),
        "最大Na连通分量占比": round_or_none(largest_frac, 4),
        "Na网络维度估计": dim,
    }


def percentile_ranks(values: list[float | None]) -> list[float | None]:
    finite = [(i, v) for i, v in enumerate(values) if v is not None and not math.isnan(float(v))]
    out: list[float | None] = [None] * len(values)
    n = len(finite)
    if n == 0:
        return out
    if n == 1:
        out[finite[0][0]] = 1.0
        return out
    finite.sort(key=lambda x: float(x[1]))
    pos = 0
    while pos < n:
        end = pos + 1
        while end < n and float(finite[end][1]) == float(finite[pos][1]):
            end += 1
        avg_rank = (pos + end - 1) / 2.0
        pct = avg_rank / (n - 1)
        for k in range(pos, end):
            out[finite[k][0]] = round(float(pct), 6)
        pos = end
    return out


def _spacegroup_from_cif_text(cif_path: Path) -> str | None:
    try:
        text = cif_path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return None
    sym = num = None
    m = re.search(r"_symmetry_space_group_name_H-M\s+['\"]?([^'\"\n]+?)['\"]?\s*\n", text)
    if m:
        sym = m.group(1).strip()
    if not sym:
        m = re.search(r"_space_group_name_H-M_alt\s+['\"]?([^'\"\n]+?)['\"]?\s*\n", text)
        if m:
            sym = m.group(1).strip()
    m = re.search(r"_symmetry_Int_Tables_number\s+(\d+)", text)
    if m:
        num = m.group(1)
    if not num:
        m = re.search(r"_space_group_IT_number\s+(\d+)", text)
        if m:
            num = m.group(1)
    if sym and num:
        return f"{sym} ({num})"
    return sym or (f"#{num}" if num else None)


def compute_one(cif_path: Path, row_meta: dict[str, Any] | None = None, na_neighbor_cutoff: float = 6.0) -> dict[str, Any]:
    row_meta = row_meta or {}
    issues: list[str] = []
    rec: dict[str, Any] = {
        "合并编号": row_meta.get("合并编号") or infer_merge_id(cif_path),
        "CIF文件": str(cif_path),
        "体系分类": row_meta.get("体系分类"),
        "材料/结构名": row_meta.get("材料/结构名"),
        "电导率_mS_cm-1": row_meta.get("电导率_mS_cm-1"),
        "log10电导率": log10_or_none(row_meta.get("电导率_mS_cm-1")),
        "解析状态": "成功",
    }

    try:
        parser = CifParser(str(cif_path), occupancy_tolerance=10)
        structure = parser.parse_structures(primitive=False)[0]
        for warning in getattr(parser, "warnings", []) or []:
            text = str(warning)
            if "fractional coordinates rounded to ideal values" not in text:
                issues.append(text)
    except Exception as exc:
        rec.update({"解析状态": "失败", "问题数量": 1, "问题": f"CIF 解析失败: {exc}"})
        return rec

    try:
        try:
            sga = SpacegroupAnalyzer(structure, symprec=0.01)
            spacegroup = f"{sga.get_space_group_symbol()} ({sga.get_space_group_number()})"
        except Exception:
            spacegroup = _spacegroup_from_cif_text(cif_path)
            if spacegroup:
                spacegroup = f"{spacegroup} [from CIF header]"

        species_symbols = {str(getattr(el, "symbol", str(el))) for el in structure.composition.elements}
        anions = species_symbols & ANION_ELEMENTS
        na_sites = [(i, site, site_occ(site, MOBILE)) for i, site in enumerate(structure) if site_occ(site, MOBILE) > 1e-6]
        na_indices = [i for i, _site, _occ in na_sites]
        partial_na = [(i, site, occ) for i, site, occ in na_sites if occ < 0.999]
        total_occ = sum(sum(site_species_dict(site).values()) for site in structure)
        na_occ_sum = sum(occ for _i, _site, occ in na_sites)

        all_na_x_distances: list[float] = []
        na_s_distances: list[float] = []
        cn_values: list[int] = []
        cn_counter: Counter[int] = Counter()
        env_counter: Counter[str] = Counter()
        poly_volumes: list[float] = []
        distortions: list[float] = []

        for idx, _site, _occ in na_sites:
            shell = shell_neighbors(structure, idx, anions)
            distances = [float(n["distance"]) for n in shell]
            if not distances:
                issues.append(f"Na site {idx} has no measured Na-X shell")
                continue
            cn = len(shell)
            cn_values.append(cn)
            cn_counter[cn] += 1
            shell_symbols = sorted({n["symbol"] for n in shell})
            env_counter[f"Na-{'/'.join(shell_symbols)} {cn_to_polyhedron_name(cn)}"] += 1
            all_na_x_distances.extend(distances)
            na_s_distances.extend(float(n["distance"]) for n in shell if n["symbol"] == "S")
            vol = convex_volume([n["coords"] for n in shell])
            if vol is not None:
                poly_volumes.append(vol)
            if len(distances) > 1:
                distortions.append(pstdev(distances) / mean(distances))

        if not na_sites:
            issues.append("no Na sites found")
        if not all_na_x_distances:
            issues.append("no Na-X bonds measured")

        na_main_cn = mode_int(cn_values)
        na_radius, na_radius_source = effective_na_radius(na_main_cn)
        anion_radius, anion_radius_source = effective_anion_radius(anions)
        target_center = na_radius + anion_radius if na_radius is not None and anion_radius is not None else None
        na_x_max = max(all_na_x_distances) if all_na_x_distances else None
        a2 = na_x_max / target_center if na_x_max is not None and target_center else None  # 即 局域宽松因子
        distortion_mean = mean(distortions) if distortions else None
        na_concentration = na_occ_sum / float(structure.volume) if structure.volume else None
        distortion_times_a2 = distortion_mean * a2 if distortion_mean is not None and a2 is not None else None  # 即 畸变加权宽松比
        distortion_times_max_bond = (
            distortion_mean * na_x_max if distortion_mean is not None and na_x_max is not None else None
        )  # 即 畸变加权最长键
        distortion_over_na_concentration = (
            distortion_mean / na_concentration
            if distortion_mean is not None and na_concentration is not None and na_concentration > 0
            else None
        )  # 即 畸变密度

        s_radius = ANION_EFFECTIVE_RADII_A.get("S")
        na_s_target = na_radius + s_radius if na_radius is not None and s_radius is not None else None
        na_s_avg = mean(na_s_distances) if na_s_distances else None
        na_s_norm = na_s_avg / na_s_target if na_s_avg is not None and na_s_target else None

        network = build_na_network(structure, na_indices, na_neighbor_cutoff)

        vacancy_amount = sum(1.0 - occ for _i, _site, occ in partial_na)
        vacancy_indices = [i for i, _site, _occ in partial_na]
        na_vac_min_per_vac: list[float] = []
        connected_counts: list[int] = []
        reachable_per_na: defaultdict[int, set[int]] = defaultdict(set)
        for vi in vacancy_indices:
            per_na_distances = [(ni, float(structure.get_distance(vi, ni))) for ni in na_indices if ni != vi]
            if not per_na_distances:
                continue
            d_min = min(d for _ni, d in per_na_distances)
            na_vac_min_per_vac.append(d_min)
            connected = [ni for ni, d in per_na_distances if d <= d_min + 0.30]
            connected_counts.append(len(connected))
            for ni in connected:
                reachable_per_na[ni].add(vi)

        backbone_symbols = sorted(sym for sym in species_symbols if sym != MOBILE and sym not in ANION_ELEMENTS)
        backbone_counter: Counter[str] = Counter()
        for idx, site in enumerate(structure):
            center = major_symbol(site)
            if center not in backbone_symbols:
                continue
            shell = shell_neighbors(structure, idx, anions)
            if shell:
                label = "/".join(sorted({n["symbol"] for n in shell})) or "X"
                backbone_counter[f"{center}-{label}{len(shell)}"] += 1

        rec.update(
            {
                "约化式": structure.composition.reduced_formula,
                "空间群": spacegroup,
                "结构体积_A3": round_or_none(float(structure.volume), 4),
                "Na位点数": len(na_sites),
                "Na占位总和": round_or_none(na_occ_sum, 4),
                "Na浓度_A-3": round_or_none(na_concentration, 6),
                "是否部分占位/无序": "是" if any(is_disordered_or_partial(site) for site in structure) else "否",
                "Na部分占位位点数": len(partial_na),
                "Na空位比例_按部分占位估计": round_or_none(
                    (1.0 - na_occ_sum / len(na_sites)) if na_sites else None, 5
                ),
                "Na邻近阴离子类型": "/".join(sorted(anions)) if anions else None,
                "Na-X平均键长_A": round_or_none(mean(all_na_x_distances) if all_na_x_distances else None, 4),
                "Na-X最短键长_A": round_or_none(min(all_na_x_distances) if all_na_x_distances else None, 4),
                "Na-X最长键长_A": round_or_none(na_x_max, 4),
                "Na-X键长范围_A": fmt_range(all_na_x_distances),
                "Na主配位数": na_main_cn,
                "Na配位数分布": counter_text(cn_counter),
                "Na配位环境类型分布": counter_text(env_counter),
                "Na多面体平均体积_A3": round_or_none(mean(poly_volumes) if poly_volumes else None, 4),
                "Na多面体畸变均值": round_or_none(distortion_mean, 5),
                "有效Na半径_A": round_or_none(na_radius, 4),
                "有效Na半径来源": na_radius_source,
                "有效阴离子半径_A": round_or_none(anion_radius, 4),
                "有效阴离子半径来源": anion_radius_source,
                "Na-X目标键长中心_A": round_or_none(target_center, 4),
                "局域宽松因子_A2": round_or_none(a2, 6),
                "畸变乘键长比_A5": round_or_none(distortion_times_a2, 6),
                "畸变乘最长键长_A4": round_or_none(distortion_times_max_bond, 6),
                "畸变除Na浓度_A9": round_or_none(distortion_over_na_concentration, 6),
                "Na-S平均键长_A": round_or_none(na_s_avg, 4),
                "Na-S_avg_bond_norm": round_or_none(na_s_norm, 6),
                "候选空位数": len(partial_na),
                "空位总量_按占位估计": round_or_none(vacancy_amount, 4),
                "Na-空位最近距离_A": round_or_none(min(na_vac_min_per_vac) if na_vac_min_per_vac else None, 4),
                "Na-空位平均最近邻距离_A": round_or_none(mean(na_vac_min_per_vac) if na_vac_min_per_vac else None, 4),
                "每个空位平均连接Na数": round_or_none(mean(connected_counts) if connected_counts else None, 3),
                "每个Na平均可达空位数": round_or_none(
                    mean(len(s) for s in reachable_per_na.values()) if reachable_per_na else
                    (0.0 if (na_indices and vacancy_indices) else None),
                    3,
                ),
                "主骨架阳离子类型": "/".join(backbone_symbols) if backbone_symbols else None,
                "主骨架多面体": counter_text(backbone_counter),
            }
        )
        rec.update(network)
    except Exception as exc:
        issues.append(f"描述符计算失败: {exc}")
        rec["解析状态"] = "失败"

    rec["问题数量"] = len(issues)
    rec["问题"] = "; ".join(issues) if issues else None
    return rec


def infer_merge_id(cif_path: Path) -> str:
    m = re.match(r"(MAT-\d+)", cif_path.name)
    if m:
        return m.group(1)
    return cif_path.stem


def find_cif_for_merge_id(cif_dir: Path, merge_id: str) -> Path | None:
    candidates = sorted(cif_dir.glob(f"{merge_id}__*.cif"))
    if candidates:
        return candidates[0]
    candidates = sorted(cif_dir.glob(f"{merge_id}.cif"))
    return candidates[0] if candidates else None


def read_workbook_rows(path: Path, sheet: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"在 {path} 中没有找到工作表 {sheet!r}；可用工作表：{wb.sheetnames}")
    ws = wb[sheet]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    if "合并编号" not in headers:
        raise SystemExit(f"在 {path}:{sheet} 中没有找到列：合并编号")
    rows: list[dict[str, Any]] = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: clean(vals[i]) if i < len(vals) else None for i in range(len(headers))}
        if row.get("合并编号"):
            rows.append(row)
    return rows


def collect_jobs(workbook: Path | None, sheet: str, cif_dir: Path) -> list[tuple[Path | None, dict[str, Any]]]:
    if workbook is not None and workbook.exists():
        jobs: list[tuple[Path | None, dict[str, Any]]] = []
        for row in read_workbook_rows(workbook, sheet):
            mid = str(row["合并编号"])
            jobs.append((find_cif_for_merge_id(cif_dir, mid), row))
        return jobs

    jobs = []
    for cif_path in sorted(cif_dir.glob("*.cif")):
        jobs.append((cif_path, {"合并编号": infer_merge_id(cif_path)}))
    return jobs


def finalize_batch_descriptors(records: list[dict[str, Any]]) -> None:
    neighbor_ranks = percentile_ranks([safe_float(r.get("每个Na平均Na邻居数")) for r in records])
    component_ranks = percentile_ranks([safe_float(r.get("最大Na连通分量占比")) for r in records])
    for i, rec in enumerate(records):
        parts = [x for x in (neighbor_ranks[i], component_ranks[i]) if x is not None]
        nana = mean(parts) if parts else None
        a2 = safe_float(rec.get("局域宽松因子_A2"))
        rec["Na位点连通因子_NaNa综合"] = round_or_none(nana, 6)  # 即 Na位点连通因子
        rec["局域_连通协同因子_A2xNaNa"] = round_or_none(a2 * nana if a2 is not None and nana is not None else None, 6)  # 即 局域-连通协同因子


def split_descriptor_tokens(raw_values: list[str] | None) -> list[str]:
    if not raw_values:
        return ["all"]
    tokens: list[str] = []
    for raw in raw_values:
        for token in re.split(r"[,，;；\s]+", raw):
            token = token.strip()
            if token:
                tokens.append(token)
    return tokens or ["all"]


def normalize_descriptor_name(token: str) -> str | None:
    normalized_token = token.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    if normalized_token.isdigit():
        index = int(normalized_token)
        if 1 <= index <= len(DESCRIPTOR_COLUMNS):
            return DESCRIPTOR_COLUMNS[index - 1]
        return None
    if token in DESCRIPTOR_THRESHOLDS:
        return token
    return DESCRIPTOR_ALIASES.get(token.lower())


def parse_descriptor_selection(raw_values: list[str] | None) -> list[str]:
    tokens = split_descriptor_tokens(raw_values)
    if any(token.lower() in {"all", "*", "全部", "全量"} for token in tokens):
        return DESCRIPTOR_COLUMNS.copy()

    selected: list[str] = []
    unknown: list[str] = []
    for token in tokens:
        name = normalize_descriptor_name(token)
        if name is None:
            unknown.append(token)
        elif name not in selected:
            selected.append(name)

    if unknown:
        valid = "、".join(DESCRIPTOR_COLUMNS)
        aliases = "、".join(sorted(DESCRIPTOR_ALIASES))
        raise SystemExit(
            "无法识别的描述符名称或短名："
            + "、".join(unknown)
            + "\n可用编号：1-8"
            + f"\n可用完整名称：{valid}"
            + f"\n可用短名：all、全部、{aliases}"
        )
    if not selected:
        raise SystemExit("至少需要选择 1 个描述符。")
    return [name for name in DESCRIPTOR_COLUMNS if name in selected]


def selected_output_columns(selected_descriptors: list[str]) -> list[str]:
    return BASE_OUTPUT_COLUMNS + selected_descriptors + STATUS_OUTPUT_COLUMNS


def print_descriptor_list() -> None:
    # 直观名映射（仅供显示用，不改变列名字典键）
    _INTUITIVE_NAMES = {
        "局域宽松因子_A2": "局域宽松因子",
        "Na多面体畸变均值": "Na多面体畸变均值",
        "Na位点连通因子_NaNa综合": "Na位点连通因子",
        "局域_连通协同因子_A2xNaNa": "局域-连通协同因子",
        "畸变乘键长比_A5": "畸变加权宽松比",
        "畸变乘最长键长_A4": "畸变加权最长键",
        "畸变除Na浓度_A9": "畸变密度",
        "Na-S_avg_bond_norm": "硫化物键长归一化(Na-S)",
    }
    print("可计算的 8 个强相关描述符：")
    for i, name in enumerate(DESCRIPTOR_COLUMNS, start=1):
        info = DESCRIPTOR_THRESHOLDS[name]
        threshold = info.get("screening_ge")
        threshold_text = f"筛选阈值 >= {threshold}" if threshold is not None else "暂无统一筛选阈值"
        intuitive = _INTUITIVE_NAMES.get(name, name)
        print(f"{i}. {name} [即 {intuitive}]（{threshold_text}；状态：{info.get('status')}）")
    print("\n可以直接用编号选择，例如：--descriptors 1 2 3")
    print("常用短名：A2[局域宽松因子]、畸变均值、NaNa[连通因子]、A2xNaNa[局域-连通协同]、A5[畸变加权宽松比]、A4[畸变加权最长键]、A9[畸变密度]、Na-S[硫化物键长归一化]")


def resolve_input_path(path_text: str) -> Path:
    path = Path(path_text).expanduser()
    if path.is_absolute() or path.exists():
        return path
    for base in (SCRIPT_DIR, PROJECT_ROOT, PROJECT_ROOT / "dataset"):
        candidate = base / path
        if candidate.exists():
            return candidate
    return path


def write_csv(path: Path, records: list[dict[str, Any]], columns: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)


def compact_record(rec: dict[str, Any], columns: list[str]) -> dict[str, Any]:
    return {col: rec.get(col) for col in columns}


def write_issues(path: Path, records: list[dict[str, Any]]) -> None:
    rows = []
    for rec in records:
        issue_text = rec.get("问题")
        if issue_text:
            for issue in str(issue_text).split("; "):
                rows.append({"合并编号": rec.get("合并编号"), "CIF文件": rec.get("CIF文件"), "问题": issue})
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["合并编号", "CIF文件", "问题"])
        writer.writeheader()
        writer.writerows(rows)


def selected_threshold_judgment(rec: dict[str, Any], selected_descriptors: list[str]) -> tuple[bool, str]:
    if rec.get("解析状态") != "成功":
        return False, "CIF 未成功解析"

    details: list[str] = []
    all_pass = True
    for name in selected_descriptors:
        threshold = safe_float(DESCRIPTOR_THRESHOLDS.get(name, {}).get("screening_ge"))
        value = safe_float(rec.get(name))
        if threshold is None:
            all_pass = False
            details.append(f"{name}: 没有可用快导体筛选阈值")
            continue
        if value is None:
            all_pass = False
            details.append(f"{name}: 缺失，未达到 >= {threshold:g}")
            continue
        if value >= threshold:
            details.append(f"{name}: {value:g} >= {threshold:g}")
        else:
            all_pass = False
            details.append(f"{name}: {value:g} < {threshold:g}")
    return all_pass, "；".join(details)


def resolve_existing_cif_path(path_text: Any) -> Path | None:
    if not path_text:
        return None
    path = Path(str(path_text)).expanduser()
    candidates = [path] if path.is_absolute() else [Path.cwd() / path, SCRIPT_DIR / path, PROJECT_ROOT / path]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def unique_destination_path(dest_dir: Path, source: Path) -> Path:
    dest = dest_dir / source.name
    if not dest.exists():
        return dest
    for i in range(2, 10000):
        candidate = dest_dir / f"{source.stem}__{i}{source.suffix}"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"目标文件名冲突过多，无法移动：{source.name}")


def write_move_report(path: Path, rows: list[dict[str, Any]]) -> None:
    columns = ["合并编号", "原CIF文件", "新CIF文件", "所选描述符", "阈值判定详情", "移动状态"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def move_fast_cifs_by_threshold(
    records: list[dict[str, Any]], selected_descriptors: list[str], output_dir: Path
) -> dict[str, Any]:
    dest_dir = output_dir.expanduser()
    if not dest_dir.is_absolute():
        dest_dir = Path.cwd() / dest_dir
    dest_dir.mkdir(parents=True, exist_ok=True)

    report_rows: list[dict[str, Any]] = []
    candidate_count = 0
    moved_count = 0
    skipped_count = 0
    selected_text = "、".join(selected_descriptors)

    for rec in records:
        is_fast, detail = selected_threshold_judgment(rec, selected_descriptors)
        rec["所选描述符全部达快导体阈值"] = "是" if is_fast else "否"
        rec["阈值判定详情"] = detail
        if not is_fast:
            continue

        candidate_count += 1
        source = resolve_existing_cif_path(rec.get("CIF文件"))
        original_text = rec.get("CIF文件")
        if source is None:
            skipped_count += 1
            report_rows.append(
                {
                    "合并编号": rec.get("合并编号"),
                    "原CIF文件": original_text,
                    "新CIF文件": None,
                    "所选描述符": selected_text,
                    "阈值判定详情": detail,
                    "移动状态": "未移动：没有找到 CIF 文件",
                }
            )
            continue

        try:
            if source.resolve().parent == dest_dir.resolve():
                skipped_count += 1
                rec["CIF文件"] = str(source)
                status = "未移动：文件已在目标文件夹"
                new_path_text = str(source)
            else:
                dest = unique_destination_path(dest_dir, source)
                moved_path = shutil.move(str(source), str(dest))
                moved_count += 1
                rec["CIF文件"] = moved_path
                status = "已移动"
                new_path_text = moved_path
        except Exception as exc:
            skipped_count += 1
            status = f"未移动：{exc}"
            new_path_text = None

        report_rows.append(
            {
                "合并编号": rec.get("合并编号"),
                "原CIF文件": original_text,
                "新CIF文件": new_path_text,
                "所选描述符": selected_text,
                "阈值判定详情": detail,
                "移动状态": status,
            }
        )

    report_path = dest_dir / "移动记录.csv"
    write_move_report(report_path, report_rows)
    return {
        "enabled": True,
        "output_dir": str(dest_dir),
        "report_csv": str(report_path),
        "candidate_count": candidate_count,
        "moved_count": moved_count,
        "skipped_count": skipped_count,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="只从 CIF 计算 8 个强相关结构描述符，不计算 Zeo++ 或 SoftBV。")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="用于保留材料顺序和元数据的 Excel 表格")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Excel 工作表名称")
    parser.add_argument("--cif-dir", default=DEFAULT_CIF_DIR, help="CIF 文件夹")
    parser.add_argument("--output-csv", default="part1_descriptors.csv", help="描述符 CSV 输出文件")
    parser.add_argument("--output-json", default="part1_descriptors.json", help="描述符 JSON 输出文件")
    parser.add_argument("--issues-csv", default="part1_issues.csv", help="问题记录 CSV 输出文件")
    parser.add_argument("--na-neighbor-cutoff", type=float, default=6.0, help="构建 Na-Na 连通图的距离截断，单位为 A")
    parser.add_argument("--no-workbook", action="store_true", help="忽略 Excel，直接处理 --cif-dir 中的所有 *.cif")
    parser.add_argument(
        "--move-fast-cifs",
        action="store_true",
        help="手动启用：将所选描述符全部达到快导体阈值的 CIF 移动到 --fast-cif-dir",
    )
    parser.add_argument(
        "--fast-cif-dir",
        default=DEFAULT_FAST_CIF_OUTPUT_DIR,
        help="--move-fast-cifs 启用时的目标文件夹",
    )
    parser.add_argument(
        "--descriptors",
        nargs="+",
        default=DEFAULT_SELECTED_DESCRIPTORS,
        help="指定要输出的强相关描述符；可用编号、完整中文名或短名，如 1 2 3 或 A2 A4 A9；默认 all 表示全部 8 个",
    )
    parser.add_argument("--list-descriptors", action="store_true", help="列出 8 个描述符、阈值和常用短名后退出")
    args = parser.parse_args()

    if args.list_descriptors:
        print_descriptor_list()
        return

    selected_descriptors = parse_descriptor_selection(args.descriptors)
    output_columns = selected_output_columns(selected_descriptors)

    cif_dir = resolve_input_path(args.cif_dir)
    workbook = None if args.no_workbook else resolve_input_path(args.workbook)
    if not cif_dir.exists():
        raise SystemExit(f"没有找到 CIF 文件夹：{cif_dir}")
    if workbook is not None and not workbook.exists():
        print(f"警告：没有找到 Excel 表格（{workbook}），将直接处理 CIF 文件夹")
        workbook = None

    jobs = collect_jobs(workbook, args.sheet, cif_dir)
    records: list[dict[str, Any]] = []
    missing = 0
    for cif_path, row_meta in jobs:
        if cif_path is None:
            missing += 1
            records.append(
                {
                    "合并编号": row_meta.get("合并编号"),
                    "体系分类": row_meta.get("体系分类"),
                    "材料/结构名": row_meta.get("材料/结构名"),
                    "电导率_mS_cm-1": row_meta.get("电导率_mS_cm-1"),
                    "log10电导率": log10_or_none(row_meta.get("电导率_mS_cm-1")),
                    "解析状态": "失败",
                    "问题数量": 1,
                    "问题": "没有找到对应 CIF 文件",
                }
            )
            continue
        records.append(compute_one(cif_path, row_meta, na_neighbor_cutoff=args.na_neighbor_cutoff))

    finalize_batch_descriptors(records)
    move_summary: dict[str, Any] = {"enabled": False}
    if args.move_fast_cifs:
        move_summary = move_fast_cifs_by_threshold(records, selected_descriptors, Path(args.fast_cif_dir))

    write_csv(Path(args.output_csv), records, output_columns)
    write_issues(Path(args.issues_csv), records)
    Path(args.output_json).write_text(
        json.dumps(
            {
                "record_count": len(records),
                "fast_conductivity_threshold_mS_cm": FAST_CONDUCTIVITY_THRESHOLD_MS_CM,
                "selected_descriptors": selected_descriptors,
                "descriptor_thresholds": {name: DESCRIPTOR_THRESHOLDS[name] for name in selected_descriptors},
                "note": "结构解析会计算必要中间量；CSV/JSON 只输出 --descriptors 选择的强相关描述符。",
                "move_fast_cifs": move_summary,
                "records": [compact_record(r, output_columns) for r in records],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    issue_count = sum(int(r.get("问题数量") or 0) for r in records)
    ok_count = sum(1 for r in records if r.get("解析状态") == "成功")
    print(f"处理记录数：{len(records)}")
    print(f"成功解析数：{ok_count}")
    print(f"缺失 CIF 数：{missing}")
    print(f"问题数量：{issue_count}")
    print("本次输出描述符：" + "、".join(selected_descriptors))
    if args.move_fast_cifs:
        print(f"CIF 移动目标文件夹：{move_summary['output_dir']}")
        print(f"满足所选阈值的 CIF 数：{move_summary['candidate_count']}")
        print(f"已移动 CIF 数：{move_summary['moved_count']}")
        print(f"移动记录：{move_summary['report_csv']}")
    else:
        print("CIF 移动功能：未启用")
    print(f"已写出：{args.output_csv}, {args.output_json}, {args.issues_csv}")
    print("按设计排除：Zeo++ 和 SoftBV/BVSE 描述符")


if __name__ == "__main__":
    main()
