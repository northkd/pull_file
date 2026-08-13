#!/usr/bin/env python3
"""Create a softBV valence/CIF pre-review Excel table for CIF_91."""

from __future__ import annotations

import math
import multiprocessing as mp
import re
import warnings
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pymatgen.core import Composition
from pymatgen.io.cif import CifParser


REQUESTED_INPUT_DIR = Path("presentation/73加18合并整理/CIF_91")
FALLBACK_INPUT_DIR = Path("CIF_91")
OUTPUT_PATH = Path("presentation/73加18合并整理/softBV_价态预审表.xlsx")

PROJECT_OXIDATION_STATES = {
    "Na": 1,
    "O": -2,
    "S": -2,
    "Se": -2,
    "F": -1,
    "Cl": -1,
    "Br": -1,
    "I": -1,
    "P": 5,
    "Si": 4,
    "Zr": 4,
    "Hf": 4,
    "Ti": 4,
    "Sc": 3,
    "Al": 3,
    "In": 3,
    "Y": 3,
    "Mg": 2,
    "Ca": 2,
    "Sr": 2,
    "Cr": 3,
    "Fe": 3,
    "V": 3,
    "Sn": 4,
    "Sb": 5,
    "W": 6,
}

VARIABLE_VALENCE_ELEMENTS = {"Ti", "V", "Mn", "Fe", "Cr", "Sb", "Sn"}
CHARGE_TOL = 0.1
OXI_GUESS_TIMEOUT = 5

HEADERS = [
    "CIF文件名",
    "化学式",
    "pymatgen猜测价态",
    "项目采用价态",
    "电荷平衡残差",
    "是否已有氧化态",
    "是否含部分占位",
    "是否含混合占位",
    "是否含可变价元素",
    "softBV可直接计算判断",
    "处理建议",
    "备注",
]


def fmt_state(value: float | int) -> str:
    sign = "+" if value > 0 else ""
    if float(value).is_integer():
        return f"{sign}{int(value)}"
    return f"{sign}{value:g}"


def fmt_state_map(states: dict[str, float | int] | None) -> str:
    if not states:
        return ""
    return "; ".join(f"{el}{fmt_state(ox)}" for el, ox in sorted(states.items()))


def raw_has_oxidation(text: str) -> bool:
    if "_atom_type_oxidation_number" in text:
        return True
    patterns = [
        r"\b[A-Z][a-z]?\d+[+-]\b",
        r"\b[A-Z][a-z]?[+-]\d+\b",
    ]
    return any(re.search(pattern, text) for pattern in patterns)


def has_species_oxidation(structure) -> bool:
    for site in structure:
        for species in site.species:
            if getattr(species, "oxi_state", None) is not None:
                return True
    return False


def element_amounts_from_structure(structure) -> dict[str, float]:
    comp = structure.composition.element_composition
    return {el.symbol: float(amount) for el, amount in comp.items()}


def element_amounts_from_text(text: str) -> dict[str, float]:
    formula_match = re.search(
        r"_chemical_formula_sum\s+(?:'([^']+)'|\"([^\"]+)\"|([^\n]+))",
        text,
    )
    formula = ""
    if formula_match:
        formula = next(group for group in formula_match.groups() if group)
    if not formula:
        return {}
    try:
        comp = Composition(formula)
        return {el.symbol: float(amount) for el, amount in comp.element_composition.items()}
    except Exception:
        return {}


def charge_residual(amounts: dict[str, float], states: dict[str, int]) -> float | None:
    if not amounts:
        return None
    if any(el not in states for el in amounts):
        return None
    return sum(amount * states[el] for el, amount in amounts.items())


def oxi_guess_worker(amounts: dict[str, float], queue: mp.Queue) -> None:
    try:
        comp = Composition(amounts)
        guesses = comp.oxi_state_guesses()
        queue.put(("ok", guesses))
    except Exception as exc:  # noqa: BLE001
        queue.put(("err", f"{type(exc).__name__}: {exc}"))


def pymatgen_oxi_guesses(amounts: dict[str, float]) -> tuple[str, bool]:
    if not amounts:
        return "", False
    ctx = mp.get_context("fork")
    queue: mp.Queue = ctx.Queue()
    proc = ctx.Process(target=oxi_guess_worker, args=(amounts, queue))
    proc.start()
    proc.join(OXI_GUESS_TIMEOUT)
    if proc.is_alive():
        proc.terminate()
        proc.join()
        return f"猜测超时（>{OXI_GUESS_TIMEOUT}s）", False
    if queue.empty():
        return "无法猜测：无返回", False
    status, payload = queue.get()
    if status == "err":
        return f"无法猜测：{payload}", False
    guesses = payload
    if not guesses:
        return "无电荷平衡猜测", False
    shown = [fmt_state_map({str(el): ox for el, ox in guess.items()}) for guess in guesses[:3]]
    suffix = "" if len(guesses) <= 3 else f"；另有{len(guesses) - 3}组"
    return " | ".join(shown) + suffix, True


def occupancy_flags(structure) -> tuple[bool, bool]:
    partial = False
    mixed = False
    for site in structure:
        if len(site.species) > 1:
            mixed = True
        total_occ = sum(float(occ) for occ in site.species.values())
        if abs(total_occ - 1.0) > 1e-4:
            partial = True
        if any(float(occ) < 0.9999 for occ in site.species.values()):
            partial = True
    return partial, mixed


def has_crossline_occupancy(text: str) -> bool:
    prev_atom_like = False
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("_") or line == "loop_":
            prev_atom_like = False
            continue
        if prev_atom_like and re.fullmatch(r"0?\.\d+(?:[Ee][-+]?\d+)?|1\.0+", line):
            return True
        tokens = line.split()
        prev_atom_like = bool(tokens and re.fullmatch(r"[A-Z][a-z]?\w*", tokens[0]) and len(tokens) >= 5)
    return False


def classify(
    *,
    parse_error: str,
    amounts: dict[str, float],
    missing_defaults: list[str],
    residual: float | None,
    has_partial: bool,
    has_mixed: bool,
    has_variable: bool,
    cell_abnormal: bool,
    crossline_occ: bool,
    parser_warnings: list[str],
    pymatgen_has_guess: bool,
) -> tuple[str, str, str]:
    reasons: list[str] = []
    advice: list[str] = []

    if parse_error:
        reasons.append(f"pymatgen读取失败：{parse_error}")
        advice.append("先修复 CIF 格式或重新导出标准 CIF")
        return "暂不建议计算", "；".join(advice), "；".join(reasons)

    if missing_defaults:
        reasons.append("项目默认价态缺失：" + ", ".join(missing_defaults))
        advice.append("补充默认价态或人工确认价态")

    if residual is None:
        reasons.append("无法计算项目默认价态电荷残差")
    elif abs(residual) > CHARGE_TOL:
        reasons.append(f"项目默认价态电荷不平衡，残差 {residual:.4g}")
        advice.append("人工核对组成和价态")

    if not pymatgen_has_guess:
        reasons.append("pymatgen 未给出电荷平衡价态猜测")

    if cell_abnormal:
        reasons.append("晶胞明显异常或过大")
        advice.append("检查是否为人工真空/超胞，先标准化或重建结构")

    if crossline_occ:
        reasons.append("疑似 occupancy 数值跨行，CIF 格式兼容性差")
        advice.append("先规范 CIF 行格式")

    if has_mixed:
        reasons.append("含混合占位")
        advice.append("先构造有序占位模型")
    elif has_partial:
        reasons.append("含部分占位")
        advice.append("先处理部分占位或确认 softBV 忽略迁移离子后是否可接受")

    if has_variable:
        reasons.append("含可变价元素，价态需人工复核")

    if parser_warnings:
        warning_text = "；".join(parser_warnings[:3])
        if "Incorrect stoichiometry" in warning_text:
            reasons.append("pymatgen 提示 CIF 化学计量与展开结构不一致")
            advice.append("检查对称展开、占位和原胞/常规胞")

    if missing_defaults or residual is None or (residual is not None and abs(residual) > CHARGE_TOL) or cell_abnormal:
        return "暂不建议计算", "；".join(dict.fromkeys(advice)) or "人工复核后再计算", "；".join(dict.fromkeys(reasons))

    if has_mixed or has_partial or crossline_occ or any("Incorrect stoichiometry" in w for w in parser_warnings):
        return "需清理后计算", "；".join(dict.fromkeys(advice)) or "清理 CIF 后计算", "；".join(dict.fromkeys(reasons))

    if has_variable:
        return "可直接计算", "项目默认价态电荷平衡；可计算，但建议保留可变价元素复核记录", "；".join(dict.fromkeys(reasons))

    return "可直接计算", "可按项目默认价态直接用于 softBV", "价态明确，电荷平衡，未发现明显占位或格式问题"


def review_one(path: Path) -> dict[str, object]:
    text = path.read_text(encoding="utf-8", errors="replace")
    parse_error = ""
    parser_warnings: list[str] = []
    structure = None

    try:
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            parser = CifParser(str(path))
            structure = parser.parse_structures(primitive=False)[0]
        parser_warnings = [str(w.message).replace("\n", " ") for w in caught]
    except Exception as exc:
        parse_error = f"{type(exc).__name__}: {exc}"

    if structure is not None:
        amounts = element_amounts_from_structure(structure)
        formula = structure.composition.formula
        has_existing_oxi = raw_has_oxidation(text) or has_species_oxidation(structure)
        has_partial, has_mixed = occupancy_flags(structure)
        abc = structure.lattice.abc
        cell_abnormal = max(abc) > 40 or min(abc) < 3
    else:
        amounts = element_amounts_from_text(text)
        formula = ""
        has_existing_oxi = raw_has_oxidation(text)
        has_partial = " 0." in text or "\t0." in text
        has_mixed = False
        cell_abnormal = False

    elements = sorted(amounts)
    missing_defaults = [el for el in elements if el not in PROJECT_OXIDATION_STATES]
    adopted = {el: PROJECT_OXIDATION_STATES[el] for el in elements if el in PROJECT_OXIDATION_STATES}
    residual = charge_residual(amounts, PROJECT_OXIDATION_STATES)
    pmg_guess, pmg_has_guess = pymatgen_oxi_guesses(amounts)
    variable_elements = sorted(el for el in elements if el in VARIABLE_VALENCE_ELEMENTS)
    crossline_occ = has_crossline_occupancy(text)

    decision, advice, note = classify(
        parse_error=parse_error,
        amounts=amounts,
        missing_defaults=missing_defaults,
        residual=residual,
        has_partial=has_partial,
        has_mixed=has_mixed,
        has_variable=bool(variable_elements),
        cell_abnormal=cell_abnormal,
        crossline_occ=crossline_occ,
        parser_warnings=parser_warnings,
        pymatgen_has_guess=pmg_has_guess,
    )

    if parse_error:
        note = f"{note}；原始公式解析元素：{', '.join(elements) or '无'}"
    if missing_defaults:
        note = f"{note}；缺失默认价态元素：{', '.join(missing_defaults)}"
    if variable_elements:
        note = f"{note}；可变价元素：{', '.join(variable_elements)}"

    return {
        "CIF文件名": path.name,
        "化学式": formula,
        "pymatgen猜测价态": pmg_guess,
        "项目采用价态": fmt_state_map(adopted),
        "电荷平衡残差": "" if residual is None else round(residual, 6),
        "是否已有氧化态": "是" if has_existing_oxi else "否",
        "是否含部分占位": "是" if has_partial else "否",
        "是否含混合占位": "是" if has_mixed else "否",
        "是否含可变价元素": "是" if variable_elements else "否",
        "softBV可直接计算判断": decision,
        "处理建议": advice,
        "备注": note,
        "_missing_defaults": missing_defaults,
    }


def write_excel(rows: list[dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "价态预审"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fills = {
        "可直接计算": PatternFill("solid", fgColor="E2F0D9"),
        "需清理后计算": PatternFill("solid", fgColor="FFF2CC"),
        "暂不建议计算": PatternFill("solid", fgColor="F4CCCC"),
    }

    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])
        decision = row.get("softBV可直接计算判断", "")
        fill = fills.get(str(decision))
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 48,
        "B": 28,
        "C": 42,
        "D": 42,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 16,
        "J": 18,
        "K": 36,
        "L": 70,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def main() -> int:
    input_dir = REQUESTED_INPUT_DIR if REQUESTED_INPUT_DIR.exists() else FALLBACK_INPUT_DIR
    if not input_dir.exists():
        raise SystemExit(f"找不到输入目录：{REQUESTED_INPUT_DIR} 或 {FALLBACK_INPUT_DIR}")

    cif_files = sorted(input_dir.glob("*.cif"))
    rows = [review_one(path) for path in cif_files]
    write_excel(rows, OUTPUT_PATH)

    counts: dict[str, int] = {}
    missing: set[str] = set()
    manual: list[str] = []
    for row in rows:
        decision = str(row["softBV可直接计算判断"])
        counts[decision] = counts.get(decision, 0) + 1
        missing.update(row["_missing_defaults"])  # type: ignore[arg-type]
        if decision != "可直接计算":
            manual.append(str(row["CIF文件名"]))

    print(f"输入目录: {input_dir}")
    print(f"输出文件: {OUTPUT_PATH}")
    print(f"总 CIF 数: {len(rows)}")
    for key in ["可直接计算", "需清理后计算", "暂不建议计算"]:
        print(f"{key}: {counts.get(key, 0)}")
    print("缺失默认价态的元素列表:", ", ".join(sorted(missing)) if missing else "无")
    print("最需要人工检查的 CIF 清单:")
    for name in manual[:30]:
        print("-", name)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
